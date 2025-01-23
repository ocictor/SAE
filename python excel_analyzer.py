import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference, LineChart
import logging

class ExcelAnalyzer:
    def __init__(self, csv_file):
        self.csv_file = csv_file
        self.setup_logging()

    def setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(message)s'
        )

    def create_excel_report(self):
        # Lecture du CSV
        df = pd.read_csv(self.csv_file)
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        
        # Feuille 1: Données brutes
        ws_data = wb.active
        ws_data.title = "Données Brutes"
        
        # Conversion du DataFrame en Excel
        for r in pd.DataFrame(df).itertuples():
            ws_data.append(r[1:])
        
        # Feuille 2: Analyse
        ws_analysis = wb.create_sheet("Analyse")
        
        # Analyse des IPs source
        ip_counts = df['source_ip'].value_counts()
        ws_analysis.append(["Top IPs Source", "Nombre de connexions"])
        for ip, count in ip_counts.head(10).items():
            ws_analysis.append([ip, count])
        
        # Graphique des IPs
        chart = BarChart()
        data = Reference(ws_analysis, min_col=2, min_row=1, max_row=11)
        cats = Reference(ws_analysis, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = "Top 10 IPs Source"
        ws_analysis.add_chart(chart, "D1")
        
        # Analyse horaire
        hourly = df['time'].str[:2].value_counts().sort_index()
        ws_analysis.append([])
        ws_analysis.append(["Heure", "Nombre de connexions"])
        start_row = ws_analysis.max_row
        for hour, count in hourly.items():
            ws_analysis.append([hour, count])
        
        # Graphique horaire
        line_chart = LineChart()
        data = Reference(ws_analysis, min_col=2, min_row=start_row, 
                        max_row=ws_analysis.max_row)
        cats = Reference(ws_analysis, min_col=1, min_row=start_row+1, 
                        max_row=ws_analysis.max_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        line_chart.title = "Distribution Horaire du Trafic"
        ws_analysis.add_chart(line_chart, "D20")
        
        # Mise en forme
        for col in ws_data[1]:
            col.font = Font(bold=True)
            col.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", 
                                 fill_type="solid")
        
        # Sauvegarde
        wb.save('analyse_reseau.xlsx')
        logging.info("Fichier Excel créé: analyse_reseau.xlsx")

def main():
    try:
        analyzer = ExcelAnalyzer('network_data.csv')
        analyzer.create_excel_report()
    except Exception as e:
        logging.error(f"Erreur: {str(e)}")

if __name__ == "__main__":
    main()