import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
import re
import logging
from typing import Dict, Any
from datetime import datetime
import json

class NetworkAnalyzer:
    def __init__(self, input_file: str, suspicious_threshold: int = 1000):
        self.input_file = input_file
        self.data = []
        self.suspicious_threshold = suspicious_threshold
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)

    def parse_tcpdump(self):
        pattern = r'''
            (\d{2}:\d{2}:\d{2}\.\d+)\s+    # Timestamp
            IP\s+                           # IP marker
            ([\w\-\.]+?)\.?(\d+)?\s+>\s+   # Source IP and port
            ([\w\-\.]+?)\.?(\d+)?:?\s*     # Destination IP and port
            (?:Flags\s+\[(.*?)\])?         # Optional flags
            (?:\s+length\s+(\d+))?         # Optional packet length
        '''
        try:
            with open(self.input_file, 'r', encoding='utf-8') as f:
                for line in f:
                    match = re.search(pattern, line.strip(), re.VERBOSE)
                    if match:
                        entry = {
                            'timestamp': match.group(1),
                            'src_ip': match.group(2),
                            'src_port': match.group(3) or 'unknown',
                            'dst_ip': match.group(4),
                            'dst_port': match.group(5) or 'unknown',
                            'flags': match.group(6) or '',
                            'length': int(match.group(7)) if match.group(7) else 0
                        }
                        self.data.append(entry)
            self.logger.info(f"Successfully parsed {len(self.data)} entries")
        except Exception as e:
            self.logger.error(f"Error parsing file: {str(e)}")
            raise

    def analyze_traffic(self):
        if not self.data:
            return {}
            
        df = pd.DataFrame(self.data)
        src_ip_counts = df['src_ip'].value_counts()
        suspicious_ips = src_ip_counts[src_ip_counts > self.suspicious_threshold]
        dst_port_counts = df['dst_port'].value_counts()
        suspicious_ports = dst_port_counts[dst_port_counts > self.suspicious_threshold]
        
        df['hour'] = pd.to_datetime(df['timestamp'].str[:8], format='%H:%M:%S').dt.hour
        hourly_traffic = df['hour'].value_counts().sort_index()
        
        # Création des graphiques
        plt.figure(figsize=(15, 10))
        plt.subplot(2, 1, 1)
        hourly_traffic.plot(kind='line', marker='o')
        plt.title('Trafic par Heure')
        plt.xlabel('Heure')
        plt.ylabel('Nombre de Paquets')
        
        plt.subplot(2, 1, 2)
        src_ip_counts.head(10).plot(kind='bar')
        plt.title('Top 10 IPs Sources')
        plt.xlabel('IP Source')
        plt.ylabel('Nombre de Paquets')
        plt.xticks(rotation=45)
        
        plt.tight_layout()
        plt.savefig('static/traffic_analysis.png')
        plt.close()

        # Générer le CSV
        df.to_csv('network_analysis.csv', index=False)
        
        return {
            'suspicious_ips': suspicious_ips.to_dict(),
            'suspicious_ports': suspicious_ports.to_dict(),
            'hourly_traffic': hourly_traffic.to_dict(),
            'top_ips': src_ip_counts.head(10).to_dict()
        }

    def create_excel_report(self):
        # Lire le CSV généré
        df = pd.read_csv('network_analysis.csv')
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 1. Feuille de données brutes
        ws1 = wb.active
        ws1.title = 'Raw Data'
        
        # Écrire les en-têtes
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Écrire les données
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
        # 2. Feuille d'analyse des IPs
        ws2 = wb.create_sheet('IP Analysis')
        
        # En-têtes
        ws2.cell(row=1, column=1, value='IP Address').fill = header_fill
        ws2.cell(row=1, column=2, value='Packet Count').fill = header_fill
        ws2.cell(row=1, column=3, value='Total Bytes').fill = header_fill
        
        # Données
        ip_stats = df.groupby('src_ip').agg({
            'timestamp': 'count',
            'length': 'sum'
        }).reset_index()
        
        for idx, row in ip_stats.iterrows():
            ws2.cell(row=idx+2, column=1, value=row['src_ip'])
            ws2.cell(row=idx+2, column=2, value=row['timestamp'])
            ws2.cell(row=idx+2, column=3, value=row['length'])
        
        # Graphique IP Analysis
        chart = BarChart()
        chart.title = "Top IPs by Packet Count"
        chart.style = 10
        data = Reference(ws2, min_col=2, min_row=1, max_row=11)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "E2")
        
        # 3. Feuille d'analyse temporelle
        ws3 = wb.create_sheet('Time Analysis')
        
        df['hour'] = pd.to_datetime(df['timestamp'], format='%H:%M:%S.%f').dt.hour
        time_stats = df.groupby('hour').agg({
            'timestamp': 'count',
            'length': ['sum', 'mean']
        }).round(2)
        
        ws3.cell(row=1, column=1, value='Hour').fill = header_fill
        ws3.cell(row=1, column=2, value='Packet Count').fill = header_fill
        ws3.cell(row=1, column=3, value='Total Bytes').fill = header_fill
        ws3.cell(row=1, column=4, value='Average Packet Size').fill = header_fill
        
        for idx, (hour, data) in enumerate(time_stats.iterrows()):
            ws3.cell(row=idx+2, column=1, value=hour)
            ws3.cell(row=idx+2, column=2, value=data[('timestamp', 'count')])
            ws3.cell(row=idx+2, column=3, value=data[('length', 'sum')])
            ws3.cell(row=idx+2, column=4, value=data[('length', 'mean')])
        
        # Graphique temporel
        line = LineChart()
        line.title = "Traffic Distribution Over Time"
        line.style = 12
        line.y_axis.title = 'Packet Count'
        line.x_axis.title = 'Hour'
        
        data = Reference(ws3, min_col=2, min_row=1, max_row=25)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=25)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ws3.add_chart(line, "F2")
        
        # Ajuster les largeurs de colonnes
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        
        # Sauvegarder
        wb.save('network_analysis.xlsx')
        self.logger.info("Excel report generated: network_analysis.xlsx")

def main():
    try:
        # Créer le dossier static s'il n'existe pas
        import os
        if not os.path.exists('static'):
            os.makedirs('static')
            
        analyzer = NetworkAnalyzer('DumpFile.txt')
        analyzer.parse_tcpdump()
        results = analyzer.analyze_traffic()
        analyzer.create_excel_report()
        print("Analyse terminée. Les fichiers suivants ont été générés:")
        print("- network_analysis.csv")
        print("- network_analysis.xlsx")
        print("- static/traffic_analysis.png")
    except Exception as e:
        logging.error(f"Erreur lors de l'analyse: {str(e)}")
        raise

if __name__ == "__main__":
    main()